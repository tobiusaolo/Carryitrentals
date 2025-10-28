from fastapi import APIRouter, Depends, HTTPException, status, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional, Dict
from datetime import date, datetime, timedelta
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ..database import get_db
from ..auth import get_current_active_user, require_roles
from ..schemas.analytics import PropertyAnalytics, PaymentAnalytics, MaintenanceAnalytics
from ..services.analytics import analytics_service
from ..models.user import User
from ..models.property import Property
from ..models.unit import Unit
from ..models.tenant import Tenant
from ..models.payment import Payment
from ..models.enums import PaymentStatus, UnitStatus

router = APIRouter(prefix="/analytics", tags=["Analytics & Reports"])

@router.get("/owner-financial-summary", response_model=Dict)
async def get_owner_financial_summary(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get detailed financial summary for owner including per-property breakdown"""
    try:
        from datetime import date as date_class
        from decimal import Decimal
        
        # Get properties based on user role
        if current_user.role == "admin":
            # Admins see all properties
            owner_properties = db.query(Property).all()
        else:
            # Owners see only their properties
            owner_properties = db.query(Property).filter(Property.owner_id == current_user.id).all()
        
        property_ids = [p.id for p in owner_properties]
        
        if not property_ids:
            return {"properties": [], "overall": {
                "expected_monthly_revenue": 0,
                "current_month_collected": 0,
                "remaining_to_collect": 0,
                "total_tenants": 0,
                "tenants_paid": 0,
                "tenants_unpaid": 0,
                "tenants_paid_ahead": 0,
                "total_units": 0,
                "occupied_units": 0,
                "vacant_units": 0,
                "occupancy_rate": 0,
                "collection_rate": 0
            }, "current_month": date_class.today().strftime("%B %Y")}
        
        # Current month dates
        today = date_class.today()
        current_month_start = today.replace(day=1)
        
        overall_stats = {
            "expected_monthly_revenue": 0,
            "current_month_collected": 0,
            "remaining_to_collect": 0,
            "total_tenants": 0,
            "tenants_paid": 0,
            "tenants_unpaid": 0,
            "tenants_paid_ahead": 0,
            "total_units": 0,
            "occupied_units": 0,
            "vacant_units": 0
        }
        
        property_details = []
        
        for property in owner_properties:
            # Get all units for this property
            units = db.query(Unit).filter(Unit.property_id == property.id).all()
            occupied_units = [u for u in units if u.status == UnitStatus.OCCUPIED]
            vacant_units = [u for u in units if u.status == UnitStatus.AVAILABLE]
            
            # Get all tenants for this property
            tenants = db.query(Tenant).filter(
                Tenant.property_id == property.id,
                Tenant.is_active == True
            ).all()
            
            # Expected revenue = sum of monthly rent for all active tenants
            expected_revenue = sum(float(t.monthly_rent) for t in tenants)
            
            # Get payments for this month for this property's tenants
            tenant_ids = [t.id for t in tenants]
            current_month_payments = db.query(Payment).filter(
                Payment.tenant_id.in_(tenant_ids),
                Payment.paid_date >= current_month_start,
                Payment.payment_type == "rent"
            ).all()
            
            collected_this_month = sum(
                float(p.amount) for p in current_month_payments 
                if p.status == PaymentStatus.PAID
            )
            
            # Analyze tenant payment status
            tenants_paid = []
            tenants_unpaid = []
            tenants_paid_ahead = []
            
            for tenant in tenants:
                # Get tenant's payments this month
                tenant_payments = [p for p in current_month_payments if p.tenant_id == tenant.id]
                total_paid = sum(float(p.amount) for p in tenant_payments if p.status == PaymentStatus.PAID)
                
                expected_amount = float(tenant.monthly_rent)
                
                if total_paid >= expected_amount:
                    tenants_paid.append({
                        "id": tenant.id,
                        "name": f"{tenant.first_name} {tenant.last_name}",
                        "amount_paid": total_paid,
                        "expected": expected_amount
                    })
                    
                    # Check if paid ahead (more than monthly rent)
                    if total_paid > expected_amount:
                        months_ahead = int((total_paid - expected_amount) / expected_amount)
                        if months_ahead > 0:
                            tenants_paid_ahead.append({
                                "id": tenant.id,
                                "name": f"{tenant.first_name} {tenant.last_name}",
                                "months_ahead": months_ahead,
                                "extra_amount": total_paid - expected_amount
                            })
                else:
                    tenants_unpaid.append({
                        "id": tenant.id,
                        "name": f"{tenant.first_name} {tenant.last_name}",
                        "amount_paid": total_paid,
                        "expected": expected_amount,
                        "remaining": expected_amount - total_paid,
                        "payment_status": tenant.rent_payment_status
                    })
            
            property_data = {
                "property_id": property.id,
                "property_name": property.name,
                "expected_monthly_revenue": expected_revenue,
                "current_month_collected": collected_this_month,
                "remaining_to_collect": expected_revenue - collected_this_month,
                "total_tenants": len(tenants),
                "tenants_paid": len(tenants_paid),
                "tenants_unpaid": len(tenants_unpaid),
                "tenants_paid_ahead": len(tenants_paid_ahead),
                "total_units": len(units),
                "occupied_units": len(occupied_units),
                "vacant_units": len(vacant_units),
                "occupancy_rate": (len(occupied_units) / len(units) * 100) if units else 0,
                "collection_rate": (collected_this_month / expected_revenue * 100) if expected_revenue > 0 else 0,
                "tenants_paid_list": tenants_paid,
                "tenants_unpaid_list": tenants_unpaid,
                "tenants_paid_ahead_list": tenants_paid_ahead
            }
            
            property_details.append(property_data)
            
            # Update overall stats
            overall_stats["expected_monthly_revenue"] += expected_revenue
            overall_stats["current_month_collected"] += collected_this_month
            overall_stats["total_tenants"] += len(tenants)
            overall_stats["tenants_paid"] += len(tenants_paid)
            overall_stats["tenants_unpaid"] += len(tenants_unpaid)
            overall_stats["tenants_paid_ahead"] += len(tenants_paid_ahead)
            overall_stats["total_units"] += len(units)
            overall_stats["occupied_units"] += len(occupied_units)
            overall_stats["vacant_units"] += len(vacant_units)
        
        overall_stats["remaining_to_collect"] = (
            overall_stats["expected_monthly_revenue"] - 
            overall_stats["current_month_collected"]
        )
        overall_stats["occupancy_rate"] = (
            (overall_stats["occupied_units"] / overall_stats["total_units"] * 100) 
            if overall_stats["total_units"] > 0 else 0
        )
        overall_stats["collection_rate"] = (
            (overall_stats["current_month_collected"] / overall_stats["expected_monthly_revenue"] * 100)
            if overall_stats["expected_monthly_revenue"] > 0 else 0
        )
        
        return {
            "properties": property_details,
            "overall": overall_stats,
            "current_month": today.strftime("%B %Y")
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating financial summary: {str(e)}"
        )

@router.get("/dashboard-summary", response_model=Dict)
async def get_dashboard_summary(
    current_user: User = Depends(require_roles(["owner"])),
    db: Session = Depends(get_db)
):
    """Get comprehensive dashboard summary for owners"""
    try:
        # Get owner's properties
        owner_properties = db.query(Property).filter(Property.owner_id == current_user.id).all()
        property_ids = [p.id for p in owner_properties]
        
        if not property_ids:
            return {
                "total_properties": 0,
                "total_units": 0,
                "occupied_units": 0,
                "available_units": 0,
                "total_monthly_rent": 0,
                "collected_rent": 0,
                "pending_rent": 0,
                "overdue_rent": 0,
                "occupancy_rate": 0,
                "collection_rate": 0
            }
        
        # Get all units for owner's properties
        all_units = db.query(Unit).filter(Unit.property_id.in_(property_ids)).all()
        occupied_units = [u for u in all_units if u.status == UnitStatus.OCCUPIED]
        available_units = [u for u in all_units if u.status == UnitStatus.AVAILABLE]
        
        # Calculate rent metrics
        total_monthly_rent = sum(float(u.monthly_rent) for u in occupied_units)
        
        # Get current month payments
        current_month = datetime.now().replace(day=1)
        payments_this_month = db.query(Payment).filter(
            Payment.tenant_id.in_([t.id for t in db.query(Tenant).filter(Tenant.unit_id.in_([u.id for u in occupied_units]))]),
            Payment.paid_date >= current_month,
            Payment.payment_type == "rent"
        ).all()
        
        collected_rent = sum(float(p.amount) for p in payments_this_month if p.status == PaymentStatus.PAID)
        pending_rent = sum(float(p.amount) for p in payments_this_month if p.status == PaymentStatus.PENDING)
        overdue_rent = sum(float(p.amount) for p in payments_this_month if p.status == PaymentStatus.OVERDUE)
        
        # Get utility payments for owner's properties
        utility_payments_this_month = db.query(Payment).filter(
            Payment.payment_type == "utility",
            Payment.paid_date >= current_month,
            Payment.unit_id.in_([u.id for u in all_units])
        ).all()
        
        collected_utilities = sum(float(p.amount) for p in utility_payments_this_month if p.status == PaymentStatus.PAID)
        pending_utilities = sum(float(p.amount) for p in utility_payments_this_month if p.status == PaymentStatus.PENDING)
        overdue_utilities = sum(float(p.amount) for p in utility_payments_this_month if p.status == PaymentStatus.OVERDUE)
        
        # Get total utility costs for owner's properties
        from ..models.utility import Utility
        from ..models.unit_utility import UnitUtility
        
        property_utilities = db.query(Utility).filter(Utility.property_id.in_(property_ids)).all()
        unit_utilities = db.query(UnitUtility).join(Unit).filter(Unit.property_id.in_(property_ids)).all()
        
        total_monthly_utilities = (
            sum(float(u.monthly_cost) for u in property_utilities) +
            sum(float(u.monthly_cost) for u in unit_utilities)
        )
        
        # Calculate rates
        occupancy_rate = (len(occupied_units) / len(all_units) * 100) if all_units else 0
        collection_rate = (collected_rent / total_monthly_rent * 100) if total_monthly_rent > 0 else 0
        utility_collection_rate = (collected_utilities / total_monthly_utilities * 100) if total_monthly_utilities > 0 else 0
        
        return {
            "total_properties": len(owner_properties),
            "total_units": len(all_units),
            "occupied_units": len(occupied_units),
            "available_units": len(available_units),
            "total_monthly_rent": total_monthly_rent,
            "collected_rent": collected_rent,
            "pending_rent": pending_rent,
            "overdue_rent": overdue_rent,
            "occupancy_rate": round(occupancy_rate, 2),
            "collection_rate": round(collection_rate, 2),
            "total_monthly_utilities": total_monthly_utilities,
            "collected_utilities": collected_utilities,
            "pending_utilities": pending_utilities,
            "overdue_utilities": overdue_utilities,
            "utility_collection_rate": round(utility_collection_rate, 2)
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching dashboard summary: {str(e)}"
        )

@router.get("/property/{property_id}", response_model=Dict)
async def get_property_analytics(
    property_id: int,
    current_user: User = Depends(require_roles(["owner"])),
    db: Session = Depends(get_db)
):
    """Get detailed analytics for a specific property"""
    try:
        # Check property access
        property = db.query(Property).filter(Property.id == property_id).first()
        if not property:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Property not found"
            )
        
        if property.owner_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
        
        # Get property units
        units = db.query(Unit).filter(Unit.property_id == property_id).all()
        occupied_units = [u for u in units if u.status == UnitStatus.OCCUPIED]
        
        # Get tenants for occupied units
        tenants = db.query(Tenant).filter(Tenant.unit_id.in_([u.id for u in occupied_units])).all()
        
        # Get payments for tenants
        tenant_ids = [t.id for t in tenants]
        payments = db.query(Payment).filter(Payment.tenant_id.in_(tenant_ids)).all()
        
        # Calculate metrics
        total_monthly_rent = sum(float(u.monthly_rent) for u in occupied_units)
        current_month = datetime.now().replace(day=1)
        current_month_payments = [p for p in payments if p.payment_date >= current_month]
        
        collected_rent = sum(float(p.amount) for p in current_month_payments if p.status == PaymentStatus.COMPLETED)
        pending_rent = sum(float(p.amount) for p in current_month_payments if p.status == PaymentStatus.PENDING)
        overdue_rent = sum(float(p.amount) for p in current_month_payments if p.status == PaymentStatus.OVERDUE)
        
        occupancy_rate = (len(occupied_units) / len(units) * 100) if units else 0
        collection_rate = (collected_rent / total_monthly_rent * 100) if total_monthly_rent > 0 else 0
        
        return {
            "property_id": property_id,
            "property_name": property.name,
            "property_address": f"{property.address}, {property.city}",
            "total_units": len(units),
            "occupied_units": len(occupied_units),
            "available_units": len(units) - len(occupied_units),
            "total_monthly_rent": total_monthly_rent,
            "collected_rent": collected_rent,
            "pending_rent": pending_rent,
            "overdue_rent": overdue_rent,
            "occupancy_rate": round(occupancy_rate, 2),
            "collection_rate": round(collection_rate, 2),
            "units": [
                {
                    "unit_id": u.id,
                    "unit_number": u.unit_number,
                    "status": u.status.value,
                    "monthly_rent": float(u.monthly_rent),
                    "bedrooms": u.bedrooms,
                    "bathrooms": u.bathrooms,
                    "tenant_name": next((f"{t.first_name} {t.last_name}" for t in tenants if t.unit_id == u.id), None)
                }
                for u in units
            ]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching property analytics: {str(e)}"
        )

@router.get("/payments", response_model=Dict)
async def get_payment_analytics(
    property_id: Optional[int] = None,
    unit_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(require_roles(["owner"])),
    db: Session = Depends(get_db)
):
    """Get comprehensive payment analytics with filtering"""
    try:
        # Get owner's properties
        owner_properties = db.query(Property).filter(Property.owner_id == current_user.id).all()
        property_ids = [p.id for p in owner_properties]
        
        if not property_ids:
            return {"payments": [], "summary": {}}
        
        # Build query filters
        query_filters = []
        
        if property_id:
            if property_id not in property_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Access denied to this property"
                )
            query_filters.append(Property.id == property_id)
        
        if unit_id:
            unit = db.query(Unit).filter(Unit.id == unit_id).first()
            if not unit or unit.property_id not in property_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Access denied to this unit"
                )
            query_filters.append(Unit.id == unit_id)
        
        # Get payments with filters
        payments_query = db.query(Payment).join(Tenant).join(Unit).join(Property).filter(
            Property.id.in_(property_ids)
        )
        
        if query_filters:
            payments_query = payments_query.filter(*query_filters)
        
        if start_date:
            payments_query = payments_query.filter(Payment.paid_date >= start_date)
        
        if end_date:
            payments_query = payments_query.filter(Payment.paid_date <= end_date)
        
        payments = payments_query.all()
        
        # Calculate summary
        total_amount = sum(float(p.amount) for p in payments)
        completed_amount = sum(float(p.amount) for p in payments if p.status == PaymentStatus.COMPLETED)
        pending_amount = sum(float(p.amount) for p in payments if p.status == PaymentStatus.PENDING)
        overdue_amount = sum(float(p.amount) for p in payments if p.status == PaymentStatus.OVERDUE)
        
        # Format payment data
        payment_data = []
        for payment in payments:
            tenant = db.query(Tenant).filter(Tenant.id == payment.tenant_id).first()
            unit = db.query(Unit).filter(Unit.id == tenant.unit_id).first() if tenant else None
            property_obj = db.query(Property).filter(Property.id == unit.property_id).first() if unit else None
            
            payment_data.append({
                "payment_id": payment.id,
                "tenant_name": f"{tenant.first_name} {tenant.last_name}" if tenant else "Unknown",
                "unit_number": unit.unit_number if unit else "Unknown",
                "property_name": property_obj.name if property_obj else "Unknown",
                "amount": float(payment.amount),
                "payment_date": payment.paid_date.isoformat() if payment.paid_date else None,
                "status": payment.status.value,
                "payment_type": payment.payment_type.value,
                "description": payment.description
            })
        
        return {
            "payments": payment_data,
            "summary": {
                "total_payments": len(payments),
                "total_amount": total_amount,
                "completed_amount": completed_amount,
                "pending_amount": pending_amount,
                "overdue_amount": overdue_amount,
                "completion_rate": round((completed_amount / total_amount * 100) if total_amount > 0 else 0, 2)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching payment analytics: {str(e)}"
        )

@router.get("/export/excel")
async def export_analytics_to_excel(
    export_type: str = "all",  # all, properties, units, payments, tenants
    property_id: Optional[int] = None,
    unit_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(require_roles(["owner"])),
    db: Session = Depends(get_db)
):
    """Export analytics data to Excel with comprehensive filtering"""
    try:
        # Get owner's properties
        owner_properties = db.query(Property).filter(Property.owner_id == current_user.id).all()
        property_ids = [p.id for p in owner_properties]
        
        if not property_ids:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No properties found for this owner"
            )
        
        # Create Excel workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if export_type in ["all", "properties"]:
            # Properties Sheet
            ws_properties = wb.create_sheet("Properties")
            
            # Properties data
            properties_data = []
            for prop in owner_properties:
                units = db.query(Unit).filter(Unit.property_id == prop.id).all()
                occupied_units = [u for u in units if u.status == UnitStatus.OCCUPIED]
                total_rent = sum(float(u.monthly_rent) for u in occupied_units)
                
                properties_data.append({
                    "Property ID": prop.id,
                    "Property Name": prop.name,
                    "Address": f"{prop.address}, {prop.city}",
                    "Total Units": len(units),
                    "Occupied Units": len(occupied_units),
                    "Available Units": len(units) - len(occupied_units),
                    "Occupancy Rate": f"{(len(occupied_units) / len(units) * 100) if units else 0:.1f}%",
                    "Total Monthly Rent": total_rent,
                    "Created Date": prop.created_at.strftime("%Y-%m-%d") if prop.created_at else "N/A"
                })
            
            # Write properties data
            if properties_data:
                df_properties = pd.DataFrame(properties_data)
                for r in dataframe_to_rows(df_properties, index=False, header=True):
                    ws_properties.append(r)
                
                # Style headers
                for cell in ws_properties[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Auto-adjust column widths
                for column in ws_properties.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_properties.column_dimensions[column_letter].width = adjusted_width
        
        if export_type in ["all", "units"]:
            # Units Sheet
            ws_units = wb.create_sheet("Units")
            
            # Units data
            units_query = db.query(Unit).join(Property).filter(Property.id.in_(property_ids))
            if property_id:
                units_query = units_query.filter(Unit.property_id == property_id)
            
            units = units_query.all()
            units_data = []
            
            for unit in units:
                property_obj = db.query(Property).filter(Property.id == unit.property_id).first()
                tenant = db.query(Tenant).filter(Tenant.unit_id == unit.id).first()
                
                units_data.append({
                    "Unit ID": unit.id,
                    "Unit Number": unit.unit_number,
                    "Property Name": property_obj.name if property_obj else "Unknown",
                    "Status": unit.status.value,
                    "Monthly Rent": float(unit.monthly_rent),
                    "Bedrooms": unit.bedrooms,
                    "Bathrooms": unit.bathrooms,
                    "Unit Type": unit.unit_type.value,
                    "Floor": unit.floor,
                    "Tenant Name": f"{tenant.first_name} {tenant.last_name}" if tenant else "Available",
                    "Created Date": unit.created_at.strftime("%Y-%m-%d") if unit.created_at else "N/A"
                })
            
            # Write units data
            if units_data:
                df_units = pd.DataFrame(units_data)
                for r in dataframe_to_rows(df_units, index=False, header=True):
                    ws_units.append(r)
                
                # Style headers
                for cell in ws_units[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Auto-adjust column widths
                for column in ws_units.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_units.column_dimensions[column_letter].width = adjusted_width
        
        if export_type in ["all", "payments"]:
            # Payments Sheet
            ws_payments = wb.create_sheet("Payments")
            
            # Payments data with filters
            payments_query = db.query(Payment).join(Tenant).join(Unit).join(Property).filter(
                Property.id.in_(property_ids)
            )
            
            if property_id:
                payments_query = payments_query.filter(Property.id == property_id)
            
            if unit_id:
                payments_query = payments_query.filter(Unit.id == unit_id)
            
            if start_date:
                payments_query = payments_query.filter(Payment.paid_date >= start_date)
            
            if end_date:
                payments_query = payments_query.filter(Payment.paid_date <= end_date)
            
            payments = payments_query.all()
            payments_data = []
            
            for payment in payments:
                tenant = db.query(Tenant).filter(Tenant.id == payment.tenant_id).first()
                unit = db.query(Unit).filter(Unit.id == tenant.unit_id).first() if tenant else None
                property_obj = db.query(Property).filter(Property.id == unit.property_id).first() if unit else None
                
                payments_data.append({
                    "Payment ID": payment.id,
                    "Tenant Name": f"{tenant.first_name} {tenant.last_name}" if tenant else "Unknown",
                    "Unit Number": unit.unit_number if unit else "Unknown",
                    "Property Name": property_obj.name if property_obj else "Unknown",
                    "Amount": float(payment.amount),
                    "Payment Date": payment.paid_date.strftime("%Y-%m-%d") if payment.paid_date else "N/A",
                    "Status": payment.status.value,
                    "Payment Type": payment.payment_type.value,
                    "Description": payment.description or "N/A"
                })
            
            # Write payments data
            if payments_data:
                df_payments = pd.DataFrame(payments_data)
                for r in dataframe_to_rows(df_payments, index=False, header=True):
                    ws_payments.append(r)
                
                # Style headers
                for cell in ws_payments[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Auto-adjust column widths
                for column in ws_payments.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_payments.column_dimensions[column_letter].width = adjusted_width
        
        if export_type in ["all", "tenants"]:
            # Tenants Sheet
            ws_tenants = wb.create_sheet("Tenants")
            
            # Tenants data
            tenants_query = db.query(Tenant).join(Unit).join(Property).filter(Property.id.in_(property_ids))
            if property_id:
                tenants_query = tenants_query.filter(Property.id == property_id)
            
            tenants = tenants_query.all()
            tenants_data = []
            
            for tenant in tenants:
                unit = db.query(Unit).filter(Unit.id == tenant.unit_id).first()
                property_obj = db.query(Property).filter(Property.id == unit.property_id).first() if unit else None
                
                tenants_data.append({
                    "Tenant ID": tenant.id,
                    "First Name": tenant.first_name,
                    "Last Name": tenant.last_name,
                    "Email": tenant.email or "N/A",
                    "Phone": tenant.phone or "N/A",
                    "Unit Number": unit.unit_number if unit else "Unknown",
                    "Property Name": property_obj.name if property_obj else "Unknown",
                    "Monthly Rent": float(tenant.monthly_rent),
                    "Move In Date": tenant.move_in_date.strftime("%Y-%m-%d") if tenant.move_in_date else "N/A",
                    "Move Out Date": tenant.move_out_date.strftime("%Y-%m-%d") if tenant.move_out_date else "N/A",
                    "Rent Payment Status": tenant.rent_payment_status.value,
                    "Next Payment Due": tenant.next_payment_due.strftime("%Y-%m-%d") if tenant.next_payment_due else "N/A",
                    "Last Payment Date": tenant.last_payment_date.strftime("%Y-%m-%d") if tenant.last_payment_date else "N/A"
                })
            
            # Write tenants data
            if tenants_data:
                df_tenants = pd.DataFrame(tenants_data)
                for r in dataframe_to_rows(df_tenants, index=False, header=True):
                    ws_tenants.append(r)
                
                # Style headers
                for cell in ws_tenants[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Auto-adjust column widths
                for column in ws_tenants.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_tenants.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"rental_analytics_{export_type}_{timestamp}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error exporting data: {str(e)}"
        )

@router.get("/monthly-report/{year}/{month}")
async def get_monthly_report(
    year: int,
    month: int,
    property_id: Optional[int] = None,
    current_user: User = Depends(require_roles(["owner"])),
    db: Session = Depends(get_db)
):
    """Get detailed monthly report for a specific month"""
    try:
        # Validate date
        try:
            report_date = datetime(year, month, 1)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid year or month"
            )
        
        # Get owner's properties
        owner_properties = db.query(Property).filter(Property.owner_id == current_user.id).all()
        property_ids = [p.id for p in owner_properties]
        
        if not property_ids:
            return {"message": "No properties found"}
        
        # Filter by property if specified
        if property_id:
            if property_id not in property_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Access denied to this property"
                )
            property_ids = [property_id]
        
        # Get month start and end dates
        month_start = report_date
        if month == 12:
            month_end = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = datetime(year, month + 1, 1) - timedelta(days=1)
        
        # Get units for the month
        units = db.query(Unit).join(Property).filter(Property.id.in_(property_ids)).all()
        
        # Get payments for the month
        payments = db.query(Payment).join(Tenant).join(Unit).join(Property).filter(
            Property.id.in_(property_ids),
            Payment.paid_date >= month_start,
            Payment.paid_date <= month_end
        ).all()
        
        # Calculate metrics
        occupied_units = [u for u in units if u.status == UnitStatus.OCCUPIED]
        total_monthly_rent = sum(float(u.monthly_rent) for u in occupied_units)
        
        collected_rent = sum(float(p.amount) for p in payments if p.status == PaymentStatus.COMPLETED)
        pending_rent = sum(float(p.amount) for p in payments if p.status == PaymentStatus.PENDING)
        overdue_rent = sum(float(p.amount) for p in payments if p.status == PaymentStatus.OVERDUE)
        
        return {
            "report_period": f"{year}-{month:02d}",
            "month_name": report_date.strftime("%B %Y"),
            "total_properties": len(property_ids),
            "total_units": len(units),
            "occupied_units": len(occupied_units),
            "available_units": len(units) - len(occupied_units),
            "total_monthly_rent": total_monthly_rent,
            "collected_rent": collected_rent,
            "pending_rent": pending_rent,
            "overdue_rent": overdue_rent,
            "occupancy_rate": round((len(occupied_units) / len(units) * 100) if units else 0, 2),
            "collection_rate": round((collected_rent / total_monthly_rent * 100) if total_monthly_rent > 0 else 0, 2),
            "payment_details": [
                {
                    "payment_id": p.id,
                    "tenant_name": f"{db.query(Tenant).filter(Tenant.id == p.tenant_id).first().first_name} {db.query(Tenant).filter(Tenant.id == p.tenant_id).first().last_name}",
                    "unit_number": db.query(Unit).filter(Unit.id == db.query(Tenant).filter(Tenant.id == p.tenant_id).first().unit_id).first().unit_number,
                    "amount": float(p.amount),
                    "payment_date": p.paid_date.strftime("%Y-%m-%d") if p.paid_date else None,
                    "status": p.status.value
                }
                for p in payments
            ]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating monthly report: {str(e)}"
        )